"""
JB Academy Election Portal - Utilities
Password generation, Excel import/export, results download.
"""

import random
import string
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


class Utils:

    # ── password helpers ───────────────────────────────────────────────────────

    @staticmethod
    def generate_password(length: int = 8) -> str:
        """
        Generate a memorable password: 4 letters + 4 digits.
        Avoids ambiguous chars (0/O, 1/l/I).
        """
        chars = [c for c in string.ascii_uppercase if c not in "OIL"]
        digits = [c for c in string.digits if c not in "01"]
        letters = random.choices(chars, k=4)
        nums = random.choices(digits, k=4)
        pwd = letters + nums
        random.shuffle(pwd)
        return "".join(pwd)

    @staticmethod
    def create_password_file(students: List[Tuple]) -> bytes:
        """Excel with student credentials — one sheet, sorted by class."""
        if not students:
            return b""
        data = []
        for s in students:
            data.append(
                {
                    "Admission No": s[0].upper(),
                    "Name": s[1],
                    "Class": s[2],
                    "Section": s[3],
                    "House": s[4],
                    "Password": s[6] if len(s) > 6 and s[6] else "N/A",
                    "Has Voted": "Yes" if s[7] == 1 else "No",
                }
            )
        df = pd.DataFrame(data).sort_values(["Class", "Name"])
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Student Passwords")
            ws = writer.sheets["Student Passwords"]
            _style_header_and_data(ws, df, header_color="2E5C8A")
        return output.getvalue()

    # ── results download ───────────────────────────────────────────────────────

    @staticmethod
    def create_results_file(results: Dict, stats: Dict) -> bytes:
        """
        Multi-sheet Excel:  Summary | per-committee | Participation
        Respects max_winners per committee.
        """
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # Summary
            summary_rows = []
            for ctype in ["School", "House"]:
                if ctype not in results:
                    continue
                for cname, data in sorted(results[ctype].items()):
                    cands = data["candidates"]
                    max_w = data.get("max_winners", 1)
                    winners = [c for c in cands if c.get("is_winner")][:max_w]
                    if not winners:
                        continue
                    for w in winners:
                        summary_rows.append(
                            {
                                "Type": ctype,
                                "Committee": cname,
                                "Positions": max_w,
                                "Winner": w["name"]
                                + (" (TIE)" if w.get("is_tied") else ""),
                                "Class": w["class"],
                                "House": w["house"],
                                "Votes": w["votes"],
                                "Total Votes": data["total_votes"],
                                "Vote %": f"{w['pct']}%",
                            }
                        )
            if summary_rows:
                df_sum = pd.DataFrame(summary_rows)
                df_sum.to_excel(writer, index=False, sheet_name="Summary")
                _style_header_and_data(writer.sheets["Summary"], df_sum, header_color="1F4E78")

            # Per-committee sheets
            for ctype in ["School", "House"]:
                if ctype not in results:
                    continue
                for cname, data in sorted(results[ctype].items()):
                    cands = data["candidates"]
                    if not cands:
                        continue
                    max_w = data.get("max_winners", 1)
                    rows = []
                    for rank, c in enumerate(cands, 1):
                        status = ""
                        if c.get("is_winner"):
                            status = f"{'🏆' if rank <= max_w else ''} WINNER"
                            if c.get("is_tied"):
                                status += " (TIE)"
                        rows.append(
                            {
                                "Rank": rank,
                                "Name": c["name"],
                                "Class": c["class"],
                                "House": c["house"],
                                "Votes": c["votes"],
                                "Vote %": f"{c['pct']}%",
                                "Status": status,
                            }
                        )
                    sname = f"{ctype[:1]}-{cname}"[:31]
                    df_c = pd.DataFrame(rows)
                    df_c.to_excel(writer, index=False, sheet_name=sname)
                    _style_header_and_data(writer.sheets[sname], df_c, header_color="4F81BD")

            # Participation
            part_rows = [
                {"Metric": "Phase", "Value": stats.get("phase", "N/A").upper()},
                {"Metric": "Total Students", "Value": stats["total_students"]},
                {"Metric": "Voted", "Value": stats["voted_students"]},
                {"Metric": "Not Voted", "Value": stats["not_voted"]},
                {"Metric": "Turnout %", "Value": f"{stats['participation_rate']:.1f}%"},
                {"Metric": "Total Votes", "Value": stats["total_votes"]},
            ]
            df_p = pd.DataFrame(part_rows)
            df_p.to_excel(writer, index=False, sheet_name="Participation")
            _style_header_and_data(writer.sheets["Participation"], df_p, header_color="70AD47")

        return output.getvalue()

    @staticmethod
    def create_school_results_class_wise(school_results: Dict) -> bytes:
        """
        School committee results organized by class: one sheet per class
        showing all school committees.
        Columns: Admission No, Candidate Name, Section, Committee Name, Class, Votes, Vote %
        """
        output = BytesIO()
        
        # Extract unique classes from all school committee candidates
        classes_set = set()
        for cname, data in school_results.items():
            for c in data.get("candidates", []):
                classes_set.add(c.get("class"))
        
        # Sort classes in descending order: X, IX, VIII, VII, VI, V, ...
        roman_to_int = {
            "X": 10, "IX": 9, "VIII": 8, "VII": 7, "VI": 6,
            "V": 5, "IV": 4, "III": 3, "II": 2, "I": 1
        }
        
        sorted_classes = sorted(
            classes_set,
            key=lambda c: roman_to_int.get(c, 0),
            reverse=True
        )
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for class_name in sorted_classes:
                rows = []
                
                # Iterate through all committees and collect candidates from this class
                for cname, data in sorted(school_results.items()):
                    for c in data.get("candidates", []):
                        if c.get("class") == class_name:
                            rows.append({
                                "Admission No": c["adm"],
                                "Candidate Name": c["name"],
                                "Section": c.get("section", "?"),
                                "Committee Name": cname,
                                "Class": c["class"],
                                "Votes": c["votes"],
                                "Vote %": f"{c['pct']}%",
                            })
                
                # Create sheet if there are candidates from this class
                if rows:
                    sheet_name = f"Class {class_name}"
                    df = pd.DataFrame(rows)
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    ws = writer.sheets[sheet_name]
                    
                    # Apply enhanced styling
                    _style_header_and_data(ws, df, header_color="C65911")
                    
                    # Additional formatting for school results
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    # Adjust column widths for better readability
                    ws.column_dimensions['A'].width = 12  # Admission No
                    ws.column_dimensions['B'].width = 25  # Candidate Name
                    ws.column_dimensions['C'].width = 10  # Section
                    ws.column_dimensions['D'].width = 18  # Committee Name
                    ws.column_dimensions['E'].width = 8   # Class
                    ws.column_dimensions['F'].width = 10  # Votes
                    ws.column_dimensions['G'].width = 10  # Vote %
                    
                    # Add title row at the top
                    ws.insert_rows(1)
                    ws.merge_cells('A1:G1')
                    title_cell = ws['A1']
                    title_cell.value = f"Class {class_name} - School Committee Results"
                    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                    title_cell.fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 25
                    
                    # Update freeze panes to account for title row
                    ws.freeze_panes = "A3"
        
        return output.getvalue()

    @staticmethod
    def create_house_results_file(db, results: Dict) -> bytes:
        """
        House-wise breakdown: one sheet per house showing all house committee results.
        """
        from models import Database
        
        houses = ["Ajanta", "Sanchi", "Taxila", "Nalanda"]
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for house in houses:
                rows = []
                if "House" in results:
                    for cname, data in sorted(results["House"].items()):
                        for c in data["candidates"]:
                            if (
                                c["house"] == house or True
                            ):  # show all, filter by house scope
                                pass
                        # Get candidates scoped to this house
                        # Using subquery to avoid SQLite query optimizer issue with LEFT JOIN + WHERE
                        house_cands = db.execute(
                            """
                            SELECT c.admission_no, s.name, s.class, s.section, s.house,
                                   COUNT(v.id) as vc
                            FROM (
                                SELECT * FROM candidates 
                                WHERE committee_type="House" AND scope_house=? 
                                  AND committee_name=? AND status="approved"
                            ) c
                            LEFT JOIN students s
                                ON LOWER(TRIM(c.admission_no))=LOWER(TRIM(s.admission_no))
                            LEFT JOIN votes v
                                ON LOWER(TRIM(v.candidate_adm))=LOWER(TRIM(c.admission_no))
                                AND v.committee_name=c.committee_name
                            GROUP BY c.admission_no ORDER BY vc DESC
                        """,
                            (house, cname),
                        ).fetchall()
                        if not house_cands:
                            continue
                        total = sum(r[5] for r in house_cands)
                        for rank, r in enumerate(house_cands, 1):
                            rows.append(
                                {
                                    "Committee": cname,
                                    "Rank": rank,
                                    "Admission No": r[0],
                                    "Name": r[1] or r[0],
                                    "Class": r[2] or "?",
                                    "Section": r[3] or "?",
                                    "Votes": int(r[5]),
                                    "Vote %": f"{round(r[5] / total * 100, 1) if total else 0}%",
                                }
                            )
                if rows:
                    df = pd.DataFrame(rows)
                    df.to_excel(writer, index=False, sheet_name=house[:31])
                    ws = writer.sheets[house[:31]]
                    
                    # Apply enhanced styling for house results
                    _style_header_and_data(ws, df, header_color="2E5C8A")
                    
                    # Additional formatting for house results
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    # Highlight rank 1 (winners) in each committee
                    for row_num in range(2, ws.max_row + 1):
                        rank_cell = ws[f'B{row_num}']  # Rank column
                        if rank_cell.value == 1:
                            # Gold highlight for winners
                            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                                cell = ws[f'{col_letter}{row_num}']
                                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                                cell.font = Font(bold=True, size=11)
                    
                    # Adjust column widths for better readability
                    ws.column_dimensions['A'].width = 15  # Committee
                    ws.column_dimensions['B'].width = 8   # Rank
                    ws.column_dimensions['C'].width = 12  # Admission No
                    ws.column_dimensions['D'].width = 25  # Name
                    ws.column_dimensions['E'].width = 8   # Class
                    ws.column_dimensions['F'].width = 10  # Section
                    ws.column_dimensions['G'].width = 10  # Votes
                    ws.column_dimensions['H'].width = 10  # Vote %
                    
                    # Add title row at the top
                    ws.insert_rows(1)
                    ws.merge_cells('A1:H1')
                    title_cell = ws['A1']
                    title_cell.value = f"{house} House - Committee Results"
                    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 25
                    
                    # Update freeze panes to account for title row
                    ws.freeze_panes = "A3"
                    
                else:
                    # Empty sheet placeholder
                    df = pd.DataFrame([{"Note": f"No house committee results for {house}"}])
                    df.to_excel(writer, index=False, sheet_name=house[:31])
                    _style_header_and_data(writer.sheets[house[:31]], df, header_color="2E5C8A")
        return output.getvalue()

    # ── student import ─────────────────────────────────────────────────────────

    @staticmethod
    def validate_student_data(row: dict, db) -> Tuple[bool, str, dict]:
        from models import Config, Database

        conf = Config(db)
        valid_houses = set(conf.get_houses())
        valid_classes = set(conf.get_classes())
        valid_sections = set(conf.get_sections())

        try:
            admission_no = (
                str(row.get("admission_no", "")).strip().lower().split(".")[0]
            )
            name = str(row.get("name", "")).strip()
            cls = str(row.get("class", "")).strip().split(".")[0]
            section = str(row.get("section", "A")).strip().upper()
            house = str(row.get("house", "")).strip().title()

            if not admission_no or len(admission_no) < 2:
                return False, "Invalid admission number", {}
            if not name or len(name) < 2:
                return False, "Invalid student name", {}
            if cls not in valid_classes:
                return (
                    False,
                    f"Invalid class '{cls}' (Valid: {', '.join(sorted(valid_classes))})",
                    {},
                )
            if section not in valid_sections:
                return (
                    False,
                    f"Invalid section '{section}' (Valid: {', '.join(sorted(valid_sections))})",
                    {},
                )
            if house not in valid_houses:
                return (
                    False,
                    f"Invalid house '{house}' (Valid: {', '.join(sorted(valid_houses))})",
                    {},
                )

            return (
                True,
                "",
                {
                    "admission_no": admission_no,
                    "name": name,
                    "class": cls,
                    "section": section,
                    "house": house,
                },
            )
        except Exception as e:
            return False, f"Parse error: {e}", {}

    @staticmethod
    def import_students_from_excel(
        df: pd.DataFrame, db, progress_callback=None
    ) -> Tuple[int, int, List[str]]:
        from auth import Auth
        from models import Database, Student

        imported, errors = 0, []
        student_model = Student(db)
        df.columns = [str(c).strip().lower() for c in df.columns]

        required = {"admission_no", "name", "class", "section", "house"}
        missing = required - set(df.columns)
        if missing:
            return 0, len(df), [f"Missing columns: {', '.join(sorted(missing))}"]

        total = len(df)
        
        # Pre-cache config to avoid repeated DB queries during validation
        from models import Config
        conf = Config(db)
        valid_houses = set(conf.get_houses())
        valid_classes = set(conf.get_classes())
        valid_sections = set(conf.get_sections())
        
        # Batch collect all valid records before writing to DB
        batch_inserts = []
        
        for idx, row in df.iterrows():
            if progress_callback:
                progress_callback(idx + 1, total)
            valid, err, data = Utils.validate_student_data(row.to_dict(), db)
            if not valid:
                errors.append(f"Row {idx+2}: {err}")
                continue
            if student_model.get(data["admission_no"]):
                errors.append(
                    f"Row {idx+2}: {data['admission_no']} already exists — skipped"
                )
                continue
            pwd = Utils.generate_password()
            # Queue for batch insert instead of individual writes
            batch_inserts.append({
                "admission_no": data["admission_no"],
                "name": data["name"],
                "class": data["class"],
                "section": data["section"],
                "house": data["house"],
                "pwd": pwd,
            })
        
        # Now do a single batch write to DB instead of row-by-row
        if batch_inserts:
            imported = student_model.bulk_add_batch(batch_inserts, Auth)
            if imported == 0 and batch_inserts:
                errors.append("Batch write failed — no records were imported")

        return imported, len(errors), errors

    # ── misc ───────────────────────────────────────────────────────────────────

    @staticmethod
    def format_timestamp(ts: str) -> str:
        try:
            return datetime.fromisoformat(ts).strftime("%d-%m-%Y %H:%M:%S")
        except Exception:
            return ts or ""

    @staticmethod
    def create_pending_voters_file(students: List[Tuple]) -> bytes:
        """Excel list of students who haven't voted yet."""
        if not students:
            return b""
        df = pd.DataFrame(
            students, columns=["Adm No", "Name", "Class", "Section", "House"]
        )
        df = df.sort_values(["Class", "Name"])
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Pending Voters")
            _style_header_and_data(writer.sheets["Pending Voters"], df, header_color="C55A11")
        return output.getvalue()

    @staticmethod
    def backup_results(results: Dict, stats: Dict, backup_dir: str = "backups") -> str:
        """
        Save a timestamped results Excel to disk before reset.
        Returns the filename written.
        """
        import os

        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(backup_dir, f"election_backup_{ts}.xlsx")
        data = Utils.create_results_file(results, stats)
        with open(filename, "wb") as fh:
            fh.write(data)
        return filename

    @staticmethod
    def parse_bulk_nomination_excel(df: pd.DataFrame) -> Tuple[List[Dict], List[str]]:
        """
        Parse bulk nomination upload.
        Required columns: admission_no, committee_type, committee_name
        Optional: scope_class, scope_house, section_group, manifesto
        Returns (rows, errors).
        """
        df.columns = [str(c).strip().lower() for c in df.columns]
        required = {"admission_no", "committee_type", "committee_name"}
        missing = required - set(df.columns)
        if missing:
            return [], [f"Missing columns: {', '.join(sorted(missing))}"]

        rows, errors = [], []
        valid_types = {"School", "House"}
        for idx, row in df.iterrows():
            adm = str(row.get("admission_no", "")).strip().lower().split(".")[0]
            ctype = str(row.get("committee_type", "")).strip().title()
            cname = str(row.get("committee_name", "")).strip()
            if not adm or len(adm) < 2:
                errors.append(f"Row {idx+2}: invalid admission_no")
                continue
            if ctype not in valid_types:
                errors.append(
                    f"Row {idx+2}: committee_type must be School or House, got '{ctype}'"
                )
                continue
            if not cname:
                errors.append(f"Row {idx+2}: committee_name is empty")
                continue
            rows.append(
                {
                    "admission_no": adm,
                    "committee_type": ctype,
                    "committee_name": cname,
                    "scope_class": str(row.get("scope_class", "") or "").strip()
                    or None,
                    "scope_house": str(row.get("scope_house", "") or "").strip()
                    or None,
                    "section_group": str(row.get("section_group", "") or "").strip()
                    or None,
                    "manifesto": str(row.get("manifesto", "") or "").strip(),
                }
            )
        return rows, errors

    @staticmethod
    def get_vote_statistics(db) -> dict:
        from models import Election, Database

        stats = Election(db).get_statistics()
        rows = db.execute(
            "SELECT committee_name,COUNT(*) FROM votes GROUP BY committee_name"
        ).fetchall()
        stats["votes_by_committee"] = {r[0]: r[1] for r in rows}
        return stats


def _autofit(ws, df: pd.DataFrame):
    """Auto-fit column widths and apply professional styling to worksheet."""
    # Define border style
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Style header row
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Apply header styling to first row
    for col_num, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_num)
        header_cell = ws[f'{col_letter}1']
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment
        header_cell.border = thin_border
    
    # Calculate and set column widths, apply borders and alignment to data rows
    for col_num, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_num)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None), default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
        
        # Apply styling to data rows
        for row_num, cell in enumerate(col, 1):
            if row_num > 1:  # Skip header
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Center align numeric columns (Votes, Class, etc.)
                if col_letter in ['A', 'E', 'F', 'G'] or "Votes" in str(ws[f'{col_letter}1'].value) or "Class" in str(ws[f'{col_letter}1'].value):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Alternate row colors for better readability
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Freeze header row
    ws.freeze_panes = "A2"


def _style_header_and_data(ws, df: pd.DataFrame, header_color: str = "4F81BD"):
    """Apply professional styling: headers, borders, alternating rows, auto-fit."""
    thin_border = Border(
        left=Side(style='thin', color='D0CECE'),
        right=Side(style='thin', color='D0CECE'),
        top=Side(style='thin', color='D0CECE'),
        bottom=Side(style='thin', color='D0CECE')
    )
    
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Style header row
    for col_num, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_num)
        header_cell = ws[f'{col_letter}1']
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment
        header_cell.border = thin_border
        ws.column_dimensions[col_letter].width = min(
            max((len(str(c.value)) for c in col if c.value is not None), default=10) + 3, 50
        )
    
    # Apply styling to data rows
    for row_num in range(2, ws.max_row + 1):
        for col_num, col in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{row_num}']
            cell.border = thin_border
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Center alignment for numbers and specific columns
            header_val = str(ws[f'{col_letter}1'].value or '')
            if any(x in header_val for x in ['Votes', 'Class', '%', 'No', 'Rank']):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Freeze header row
    ws.freeze_panes = "A2"
